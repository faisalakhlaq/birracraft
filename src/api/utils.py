from posixpath import split
from django.conf import settings
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils import six
from birracraft.celery import app
from api.models import *
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from datetime import datetime


def send_reset_pass_mail(request, user):
    msg=render_to_string('../templates/reset_pass_mail.html',{
        'username': user.username,
        'protocol': request.scheme,
        'domain': request.get_host(),
        'uid': urlsafe_base64_encode(force_bytes(user.pk)),
        'token': account_activation_token.make_token(user),
        })
    confirm_email = EmailMessage(
        subject='Password reset on Birracraft',
        body=msg,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[request.data['email']],
    )
    confirm_email.content_subtype = 'html'
    confirm_email.send()


def send_verification_mail(request):
    user_email = request.data['email']
    msg=render_to_string('../templates/validate_user_mail.html',{
        'username': request.data['username'],
        'protocol': request.scheme,
        'domain': request.get_host(),
        'uid': urlsafe_base64_encode(force_bytes(user_email)),
        'token': account_activation_token.make_token(user_email),
        })
    confirm_email = EmailMessage(
        subject='Verificate your user account created on Birracraft',
        body=msg,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[request.data['email']],
    )
    confirm_email.content_subtype = 'html'
    confirm_email.send()


class TokenGenerator(PasswordResetTokenGenerator):
    def _make_hash_value(self, user, timestamp):
        return (
            six.text_type(user) + six.text_type(timestamp)
        )

account_activation_token = TokenGenerator()


@app.task
def generate_report(data):

    # Get data
    orders = Order.objects.filter(
        date__gte=data['date_from']).values_list()
    orders_list = [order for order in orders]

    payments = Payment.objects.all().values_list()
    payments_list = [pay for pay in payments]
    quotas = Quota.objects.all().values_list()
    quotas_list = [quota for quota in quotas]

    products = Product.objects.all().values_list()
    products_list = [product for product in products]

    containers = Container.objects.all().values_list()
    containers_list = [container for container in containers]
    flavours = Flavour.objects.all().values_list()
    flavours_list = [flavour for flavour in flavours]

    # Make report
    wb = Workbook()
    ws = wb.active
    ws.title = 'Totals'

    ws.append(['Totals of all resources'])
    ws.append([])

    ws.append([
        'Entity',
        'Quantity'
    ])
    ws.append([
        'Orders',
        len(orders_list)
    ])
    ws.append([
        'Payments',
        len(payments_list),
    ])
    ws.append([
        'Quotas',
        len(quotas_list),
    ])
    ws.append([])
    ws.append([
        'Products',
        len(products_list)
    ])
    ws.append([
        'Containers',
        len(containers_list)
    ])
    ws.append([
        'Flavours',
        len(flavours_list)
    ])

    wb = treat_orders(wb, orders_list)
    wb = treat_payments(wb, payments_list, quotas_list)
    wb = treat_products(wb, products_list)
    wb = treat_containers_flavours(wb, containers_list, flavours_list)

    excel = save_virtual_workbook(wb)

    # Set email
    title = '[Birracraft] Report with data since {0} to {1}'.format(
                data['date_from'],
                datetime.today().strftime('%Y-%m-%d'))

    msg=render_to_string('../templates/report_mail.html',{
        'date_from': data['date_from'],
        'username': data['username'],
        })

    email = EmailMessage(
        subject=title,
        body=msg,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[data['email']]
    )
    email.content_subtype = 'html'
    email.attach(
        'Birracraft_Report_{0}.xlsx'.format(
            datetime.today().strftime('%Y-%m-%d')
        ),
        excel,
        'application/vnd.openxmlformats-officedocument\
        .spreadsheetml.sheet'
    )
    email.send()


def treat_orders(wb, orders):
    ws_orders = wb.create_sheet('Orders')
    ws_orders.append([
        'id',
        'date',
        'price',
        'delivery_cost',
        'total_amount',
        'customer',
        'payment',
        'state',
        'comment',
    ])

    for order in orders:
        ws_orders.append(order)

    return wb


def treat_payments(wb, payments, quotas):
    ws_payments = wb.create_sheet('Payments')

    # Payments
    ws_payments.append([
        'id',
        'transaction',
        'amount',
        'method',
        'quotas',
    ])

    for pay in payments:
        ws_payments.append(pay)

    ws_payments.append([])

    # Quotas
    ws_payments.append([
        'id',
        'current_quota',
        'total_quota',
        'value',
        'date',
    ])

    for quota in quotas:
        ws_payments.append(quota)

    return wb


def treat_products(wb, products):
    ws_products = wb.create_sheet('Products')
    ws_products.append([
        'id',
        'code',
        'container',
        'flavour',
        'arrived_date',
        'price',
        'state',
    ])

    for product in products:
        ws_products.append(product)

    return wb


def treat_containers_flavours(wb, containers, flavours):
    ws_c_f = wb.create_sheet('Containers_Flavours')
    ws_c_f.append([
        'id',
        'type',
        'liters',
    ])

    for container in containers:
        ws_c_f.append(container)

    ws_c_f.append([])

    ws_c_f.append([
        'id',
        'name',
        'description',
        'price_per_lt',
    ])

    for flavour in flavours:
        ws_c_f.append(flavour)

    return wb